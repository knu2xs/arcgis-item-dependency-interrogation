"""Main module for :mod:`arcgis_dependency`.

This module exposes dependency interrogation APIs used by the package root export.
"""

from collections.abc import Iterable
from pathlib import Path
from typing import Any

import pandas as pd
from arcgis.gis import GIS
from openpyxl import Workbook

from .config import config
from .utils import get_logger

logger = get_logger(__name__, level="DEBUG", add_stream_handler=False)

DEPENDENCY_COLUMNS: tuple[str, ...] = (
    "parent_item_id",
    "parent_item_name",
    "dependent_item_id",
    "dependent_item_name",
)

IMPACT_ROW_COLUMNS: tuple[str, ...] = (
    "target_item_id",
    "target_item_name",
    "affected_item_id",
    "affected_item_name",
    "relationship_status",
    "reason",
)

TARGET_DECISION_COLUMNS: tuple[str, ...] = (
    "target_item_id",
    "target_item_name",
    "decision",
    "affected_count",
    "unknown_count",
    "decision_reason",
)

ERROR_SENTINEL = "__ERROR__"


def interrogate_item_dependencies(
    item_ids: str | list[str] | None = None,
    gis: GIS | None = None,
    output_excel: str | Path | None = None,
) -> pd.DataFrame | Path:
    """Interrogate ArcGIS item dependencies for one or more item IDs.

    Args:
        item_ids: Single item ID, list of item IDs, or ``None`` to resolve from
            configuration.
        gis: Optional pre-built ArcGIS GIS instance to use instead of constructing
            one from project configuration.
        output_excel: Optional workbook path. When provided, results are written
            to Excel and the written path is returned.

    Returns:
        pandas.DataFrame | Path: Dependency table by default, or the written
        workbook path when ``output_excel`` is provided.

    Raises:
        ValueError: If no usable item IDs or GIS connection settings are found.
        RuntimeError: If item interrogation fails unexpectedly.
    """
    resolved_item_ids = _resolve_requested_item_ids(item_ids)
    if not resolved_item_ids:
        msg = "No item IDs were provided and no config request_item_ids value was available."
        logger.error(msg)
        raise ValueError(msg)

    gis = gis or _build_gis()

    rows: list[dict[str, str]] = []
    seen_edges: set[tuple[str, str]] = set()
    visited_items: set[str] = set()

    for requested_item_id in resolved_item_ids:
        try:
            item = gis.content.get(requested_item_id)
        except Exception as exc:
            msg = f"Failed to load item '{requested_item_id}' from GIS: {exc}"
            logger.warning(msg)
            rows.append(
                {
                    "parent_item_id": requested_item_id,
                    "parent_item_name": ERROR_SENTINEL,
                    "dependent_item_id": ERROR_SENTINEL,
                    "dependent_item_name": msg,
                }
            )
            continue

        if item is None:
            msg = f"Item '{requested_item_id}' was not found or is inaccessible."
            logger.warning(msg)
            rows.append(
                {
                    "parent_item_id": requested_item_id,
                    "parent_item_name": ERROR_SENTINEL,
                    "dependent_item_id": ERROR_SENTINEL,
                    "dependent_item_name": msg,
                }
            )
            continue

        visited_items.add(str(getattr(item, "id", requested_item_id)))
        _collect_dependency_rows(
            gis=gis,
            item=item,
            rows=rows,
            seen_edges=seen_edges,
            visited_items=visited_items,
        )

    dataframe = pd.DataFrame(rows, columns=DEPENDENCY_COLUMNS)
    if dataframe.empty:
        dataframe = pd.DataFrame(columns=DEPENDENCY_COLUMNS)

    if output_excel is None:
        return dataframe

    output_path = Path(output_excel)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    _write_single_sheet_workbook(output_path, dataframe, "Dependencies")
    logger.info("Wrote dependency workbook to %s", output_path)
    return output_path


def interrogate_parent_dependency_impact(
    target_item_ids: str | list[str] | None = None,
    gis: GIS | None = None,
    output_excel: str | Path | None = None,
) -> dict[str, pd.DataFrame] | Path:
    """Interrogate parent-impact by finding what items are affected by deleting targets.

    Args:
        target_item_ids: Single target ID, list of target IDs, or ``None`` to
            resolve from config ``request_item_ids``.
        gis: Optional pre-built ArcGIS GIS instance. When provided, config-based
            GIS construction is skipped.
        output_excel: Optional workbook path. When provided, writes two sheets:
            ``impact_rows`` and ``target_decisions``.

    Returns:
        dict[str, pandas.DataFrame] | Path: A dictionary containing
        ``impact_rows`` and ``target_decisions`` DataFrames by default, or the
        written workbook path when ``output_excel`` is provided.

    Raises:
        ValueError: If no usable target IDs or GIS settings are available.
        RuntimeError: If an unexpected error prevents analysis completion.
    """
    try:
        resolved_target_ids = _resolve_requested_item_ids(target_item_ids)
        if not resolved_target_ids:
            msg = (
                "No target item IDs were provided and no config request_item_ids "
                "value was available."
            )
            logger.error(msg)
            raise ValueError(msg)

        if gis is None:
            gis = _build_gis()
            logger.debug("Built GIS connection from configuration for parent-impact analysis.")
        else:
            logger.debug("Using caller-supplied GIS for parent-impact analysis.")

        all_items = _iter_accessible_items(gis)
        logger.info("Starting parent-impact analysis for %d targets across %d accessible items.", len(resolved_target_ids), len(all_items))

        impact_rows: list[dict[str, Any]] = []
        target_decisions: list[dict[str, Any]] = []
        seen_impact_edges: set[tuple[str, str]] = set()
        seen_error_rows: set[tuple[str, str, str]] = set()

        for target_item_id in resolved_target_ids:
            logger.info("Resolving target item '%s'", target_item_id)
            target_item, target_error = _resolve_item_for_target(gis, target_item_id)

            target_item_name = (
                _resolve_item_name(target_item)
                if target_item is not None
                else ERROR_SENTINEL
            )

            if target_error is not None:
                _append_error_row(
                    impact_rows=impact_rows,
                    seen_error_rows=seen_error_rows,
                    target_item_id=target_item_id,
                    target_item_name=target_item_name,
                    reason=target_error,
                )
                decision = _build_target_decision(
                    target_item_id=target_item_id,
                    target_item_name=target_item_name,
                    affected_count=0,
                    unknown_count=1,
                )
                target_decisions.append(decision)
                continue

            affected_count = 0
            unknown_count = 0

            for index, candidate_item in enumerate(all_items, start=1):
                if index % 100 == 0:
                    logger.info(
                        "Target '%s': scanned %d/%d items.",
                        target_item_id,
                        index,
                        len(all_items),
                    )

                candidate_item_id = str(getattr(candidate_item, "id", "") or "").strip()
                if not candidate_item_id or candidate_item_id == target_item_id:
                    continue

                transitive_ids, unresolved_reasons = _collect_transitive_dependency_ids(
                    gis=gis,
                    item=candidate_item,
                    visited_ids=set(),
                )

                if unresolved_reasons:
                    unknown_count += len(unresolved_reasons)
                    for unresolved_reason in unresolved_reasons:
                        _append_error_row(
                            impact_rows=impact_rows,
                            seen_error_rows=seen_error_rows,
                            target_item_id=target_item_id,
                            target_item_name=target_item_name,
                            reason=unresolved_reason,
                        )

                if target_item_id in transitive_ids:
                    edge_key = (target_item_id, candidate_item_id)
                    if edge_key in seen_impact_edges:
                        continue
                    seen_impact_edges.add(edge_key)
                    affected_count += 1
                    impact_rows.append(
                        {
                            "target_item_id": target_item_id,
                            "target_item_name": target_item_name,
                            "affected_item_id": candidate_item_id,
                            "affected_item_name": _resolve_item_name(candidate_item),
                            "relationship_status": "resolved",
                            "reason": None,
                        }
                    )

            target_decisions.append(
                _build_target_decision(
                    target_item_id=target_item_id,
                    target_item_name=target_item_name,
                    affected_count=affected_count,
                    unknown_count=unknown_count,
                )
            )

        impact_df = pd.DataFrame(impact_rows, columns=IMPACT_ROW_COLUMNS)
        if impact_df.empty:
            impact_df = pd.DataFrame(columns=IMPACT_ROW_COLUMNS)

        decisions_df = pd.DataFrame(target_decisions, columns=TARGET_DECISION_COLUMNS)
        if decisions_df.empty:
            decisions_df = pd.DataFrame(columns=TARGET_DECISION_COLUMNS)

        result = {
            "impact_rows": impact_df,
            "target_decisions": decisions_df,
        }

        if output_excel is None:
            return result

        output_path = Path(output_excel)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        _write_multi_sheet_workbook(output_path, result)
        logger.info("Wrote parent-impact workbook to %s", output_path)
        return output_path
    except ValueError:
        raise
    except Exception as exc:
        msg = f"Parent dependency impact analysis failed unexpectedly: {exc}"
        logger.error(msg)
        raise RuntimeError(msg) from exc


def _resolve_item_for_target(gis: GIS, target_item_id: str) -> tuple[Any | None, str | None]:
    """Resolve a target item by ID and return either item or an error message.

    Args:
        gis: Authenticated GIS instance.
        target_item_id: Target ID to resolve.

    Returns:
        tuple[Any | None, str | None]: Resolved item and optional error message.
    """
    try:
        target_item = gis.content.get(target_item_id)
    except Exception as exc:
        msg = f"Failed to load target '{target_item_id}' from GIS: {exc}"
        logger.warning(msg)
        return None, msg

    if target_item is None:
        msg = f"Target '{target_item_id}' was not found or is inaccessible."
        logger.warning(msg)
        return None, msg

    return target_item, None


def _iter_accessible_items(gis: GIS) -> list[Any]:
    """Return all accessible items for full-scope parent-impact discovery.

    Args:
        gis: Authenticated GIS instance.

    Returns:
        list[Any]: Item-like objects visible to the current user context.
    """
    content = getattr(gis, "content", None)
    if content is None:
        logger.warning("GIS content manager is unavailable; continuing with no accessible items.")
        return []

    try:
        items = content.search(query="*", max_items=10000)
        logger.debug("Discovered %d items using gis.content.search(query='*', max_items=10000).", len(items))
        return [item for item in items if item is not None]
    except TypeError:
        try:
            items = content.search(query="*")
            logger.debug("Discovered %d items using gis.content.search(query='*').", len(items))
            return [item for item in items if item is not None]
        except Exception as exc:
            logger.warning("Could not enumerate items using search(query='*'): %s", exc)
    except Exception as exc:
        logger.warning("Could not enumerate items using search(query='*', max_items=10000): %s", exc)

    # Fallback for local/fake GIS stubs commonly used in tests.
    items_mapping = getattr(content, "_items", None)
    if isinstance(items_mapping, dict):
        items = [item for item in items_mapping.values() if item is not None]
        logger.debug("Discovered %d items using content._items fallback.", len(items))
        return items

    logger.warning("No item enumeration strategy succeeded; returning zero accessible items.")
    return []


def _collect_transitive_dependency_ids(
    gis: GIS,
    item: Any,
    visited_ids: set[str],
) -> tuple[set[str], list[str]]:
    """Collect all recursive dependency IDs for a candidate affected item.

    Args:
        gis: Authenticated GIS instance used to fetch nested dependency items.
        item: Candidate item whose dependency closure is inspected.
        visited_ids: IDs already visited in this traversal branch.

    Returns:
        tuple[set[str], list[str]]: Transitive dependency IDs and unresolved
        dependency reasons encountered during traversal.
    """
    current_id = str(getattr(item, "id", "") or "").strip()
    if not current_id:
        return set(), ["Encountered candidate item with no id during traversal."]

    if current_id in visited_ids:
        return set(), []

    visited_ids.add(current_id)
    collected_ids: set[str] = set()
    unresolved_reasons: list[str] = []

    try:
        dependency_entries = item.get_dependencies(
            deep=False,
            outside_org=True,
            out_format="item",
        )
    except TypeError:
        dependency_entries = item.get_dependencies(deep=False, outside_org=True)
    except Exception as exc:
        msg = f"Failed to interrogate dependencies for candidate '{current_id}': {exc}"
        logger.warning(msg)
        return collected_ids, [msg]

    for dependency_entry in _normalize_dependency_entries(dependency_entries):
        dependent_id, dependent_name, dependent_item = _extract_dependency_entry(
            dependency_entry,
        )

        if not dependent_id:
            msg = f"Could not resolve a dependency ID while traversing candidate '{current_id}'."
            logger.warning(msg)
            unresolved_reasons.append(msg)
            continue

        collected_ids.add(dependent_id)

        if dependent_id in visited_ids:
            continue

        if dependent_item is None:
            try:
                dependent_item = gis.content.get(dependent_id)
            except Exception as exc:
                msg = f"Failed to resolve dependency '{dependent_id}' while traversing '{current_id}': {exc}"
                logger.warning(msg)
                unresolved_reasons.append(msg)
                continue

        if dependent_item is None:
            msg = (
                f"Dependency '{dependent_id}' ({dependent_name or dependent_id}) "
                f"was inaccessible while traversing '{current_id}'."
            )
            logger.warning(msg)
            unresolved_reasons.append(msg)
            continue

        child_ids, child_unresolved = _collect_transitive_dependency_ids(
            gis=gis,
            item=dependent_item,
            visited_ids=visited_ids,
        )
        collected_ids.update(child_ids)
        unresolved_reasons.extend(child_unresolved)

    return collected_ids, unresolved_reasons


def _append_error_row(
    impact_rows: list[dict[str, Any]],
    seen_error_rows: set[tuple[str, str, str]],
    target_item_id: str,
    target_item_name: str,
    reason: str,
) -> None:
    """Append a deduplicated error row to detailed impact output.

    Args:
        impact_rows: Mutable detailed row collection.
        seen_error_rows: Dedup set for error rows.
        target_item_id: Target ID associated with the error.
        target_item_name: Target display name.
        reason: Human-readable failure reason.
    """
    error_key = (target_item_id, ERROR_SENTINEL, reason)
    if error_key in seen_error_rows:
        return

    seen_error_rows.add(error_key)
    impact_rows.append(
        {
            "target_item_id": target_item_id,
            "target_item_name": target_item_name,
            "affected_item_id": ERROR_SENTINEL,
            "affected_item_name": ERROR_SENTINEL,
            "relationship_status": "error",
            "reason": reason,
        }
    )


def _build_target_decision(
    target_item_id: str,
    target_item_name: str,
    affected_count: int,
    unknown_count: int,
) -> dict[str, Any]:
    """Build a per-target tri-state decision record.

    Args:
        target_item_id: Target item ID.
        target_item_name: Target display name.
        affected_count: Number of resolved impacted items.
        unknown_count: Number of unresolved dependency events.

    Returns:
        dict[str, Any]: A summary decision row.
    """
    if affected_count > 0:
        decision = "not_safe_to_delete"
        reason = "Resolved affected items exist."
    elif unknown_count > 0:
        decision = "unknown_requires_review"
        reason = "Dependency resolution incomplete due to inaccessible or indeterminate data."
    else:
        decision = "safe_to_delete"
        reason = "No resolved or unresolved affected dependencies were found."

    return {
        "target_item_id": target_item_id,
        "target_item_name": target_item_name,
        "decision": decision,
        "affected_count": affected_count,
        "unknown_count": unknown_count,
        "decision_reason": reason,
    }


def _build_gis() -> GIS:
    """Build a GIS connection from project configuration."""
    gis_profile = _resolve_config_value("esri.gis_profile") or _resolve_config_value(
        "gis_profile"
    )
    gis_url = _resolve_config_value("esri.gis_url") or _resolve_config_value("gis_url")
    gis_username = _resolve_config_value("esri.gis_username") or _resolve_config_value(
        "gis_username"
    )
    gis_password = _resolve_config_value("esri.gis_password") or _resolve_config_value(
        "gis_password"
    )

    if gis_profile:
        logger.debug("Connecting to GIS using profile '%s'", gis_profile)
        return GIS(profile=str(gis_profile))

    if gis_url and gis_username and gis_password:
        logger.debug("Connecting to GIS using explicit URL and credentials.")
        return GIS(
            url=str(gis_url),
            username=str(gis_username),
            password=str(gis_password),
        )

    msg = (
        "GIS connection settings are missing. Provide esri.gis_profile or a "
        "complete URL/username/password configuration."
    )
    logger.error(msg)
    raise ValueError(msg)


def _resolve_requested_item_ids(item_ids: str | list[str] | None) -> list[str]:
    """Resolve item IDs from a direct argument or configuration fallback."""
    if item_ids is None:
        getter = getattr(config, "get", None)
        if callable(getter):
            item_ids = getter("request_item_ids", None)
            if item_ids is None:
                item_ids = getter("item_ids", None)

    return _unique_non_empty_strings(item_ids)


def _resolve_config_value(path: str) -> Any:
    """Resolve a dotted config path from the project configuration."""
    current: Any = config
    for part in path.split("."):
        if current is None:
            return None
        if isinstance(current, dict):
            current = current.get(part)
            continue
        if hasattr(current, part):
            current = getattr(current, part)
            continue
        getter = getattr(current, "get", None)
        if callable(getter):
            current = getter(part)
            continue
        return None
    return current


def _unique_non_empty_strings(value: Any) -> list[str]:
    """Normalize strings or iterables of strings into a unique ordered list."""
    if value is None:
        return []

    if isinstance(value, str):
        candidates = [value]
    elif isinstance(value, Iterable) and not isinstance(value, (bytes, bytearray)):
        candidates = list(value)
    else:
        candidates = [value]

    results: list[str] = []
    seen: set[str] = set()
    for candidate in candidates:
        item_id = str(candidate).strip()
        if not item_id or item_id in seen:
            continue
        seen.add(item_id)
        results.append(item_id)
    return results


def _collect_dependency_rows(
    gis: GIS,
    item: Any,
    rows: list[dict[str, str]],
    seen_edges: set[tuple[str, str]],
    visited_items: set[str],
) -> None:
    """Recursively collect dependency rows for an item."""
    parent_item_id = str(getattr(item, "id", "") or "").strip()
    parent_item_name = _resolve_item_name(item)

    try:
        dependency_entries = item.get_dependencies(
            deep=False,
            outside_org=True,
            out_format="item",
        )
    except TypeError:
        dependency_entries = item.get_dependencies(deep=False, outside_org=True)
    except Exception as exc:
        msg = f"Failed to interrogate dependencies for item '{parent_item_id}': {exc}"
        logger.warning(msg)
        rows.append(
            {
                "parent_item_id": parent_item_id or ERROR_SENTINEL,
                "parent_item_name": parent_item_name or ERROR_SENTINEL,
                "dependent_item_id": ERROR_SENTINEL,
                "dependent_item_name": msg,
            }
        )
        return

    for dependency_entry in _normalize_dependency_entries(dependency_entries):
        dependent_item_id, dependent_item_name, dependent_item = _extract_dependency_entry(
            dependency_entry,
        )
        if not dependent_item_id:
            msg = f"Could not resolve a dependency item for parent '{parent_item_id or ERROR_SENTINEL}'."
            logger.warning(msg)
            rows.append(
                {
                    "parent_item_id": parent_item_id or ERROR_SENTINEL,
                    "parent_item_name": parent_item_name or ERROR_SENTINEL,
                    "dependent_item_id": ERROR_SENTINEL,
                    "dependent_item_name": msg,
                }
            )
            continue

        edge_key = (parent_item_id, dependent_item_id)
        if edge_key not in seen_edges:
            rows.append(
                {
                    "parent_item_id": parent_item_id,
                    "parent_item_name": parent_item_name,
                    "dependent_item_id": dependent_item_id,
                    "dependent_item_name": dependent_item_name,
                }
            )
            seen_edges.add(edge_key)

        if dependent_item_id in visited_items:
            continue

        visited_items.add(dependent_item_id)
        if dependent_item is None:
            try:
                dependent_item = gis.content.get(dependent_item_id)
            except Exception as exc:
                logger.warning(
                    "Failed to resolve dependent item '%s' for recursion: %s",
                    dependent_item_id,
                    exc,
                )
                continue

        if dependent_item is None:
            continue

        _collect_dependency_rows(
            gis=gis,
            item=dependent_item,
            rows=rows,
            seen_edges=seen_edges,
            visited_items=visited_items,
        )


def _normalize_dependency_entries(dependency_entries: Any) -> list[Any]:
    """Normalize the raw dependency response into a list of entries."""
    if dependency_entries is None:
        return []

    if isinstance(dependency_entries, dict):
        for key in ("items", "dependencies", "results", "data"):
            if key in dependency_entries:
                dependency_entries = dependency_entries[key]
                break
        else:
            return [dependency_entries]

    if isinstance(dependency_entries, (str, bytes)):
        return [dependency_entries]

    try:
        return list(dependency_entries)
    except TypeError:
        return [dependency_entries]


def _extract_dependency_entry(entry: Any) -> tuple[str, str, Any]:
    """Extract item ID, display name, and item object from a dependency entry."""
    if entry is None:
        return "", "", None

    if isinstance(entry, str):
        item_id = entry.strip()
        return item_id, item_id, None

    if isinstance(entry, dict):
        item_id = _first_present(entry, ("id", "item_id", "dependent_item_id", "value"))
        item_name = _first_present(
            entry,
            ("title", "name", "item_name", "dependent_item_name", "displayName"),
        )
        nested_item = entry.get("item") or entry.get("dependent_item") or entry.get("value")
        if item_id is None and isinstance(nested_item, str):
            item_id = nested_item
        item_id = str(item_id).strip() if item_id is not None else ""
        item_name = str(item_name).strip() if item_name is not None else item_id
        item_obj = nested_item if hasattr(nested_item, "get_dependencies") else None
        return item_id, item_name, item_obj

    item_id = getattr(entry, "id", None) or getattr(entry, "item_id", None)
    item_name = (
        getattr(entry, "title", None)
        or getattr(entry, "name", None)
        or getattr(entry, "item_name", None)
        or getattr(entry, "dependent_item_name", None)
    )
    item_id = str(item_id).strip() if item_id is not None else ""
    item_name = str(item_name).strip() if item_name is not None else item_id
    nested_item = entry if hasattr(entry, "get_dependencies") else None
    return item_id, item_name, nested_item


def _resolve_item_name(item: Any) -> str:
    """Resolve a display name for an item-like object."""
    for attr in ("title", "name", "item_name"):
        value = getattr(item, attr, None)
        if value:
            return str(value)
    item_id = getattr(item, "id", None)
    return str(item_id) if item_id is not None else ERROR_SENTINEL


def _first_present(mapping: dict[str, Any], keys: tuple[str, ...]) -> Any:
    """Return the first present non-empty mapping value from a key sequence."""
    for key in keys:
        value = mapping.get(key)
        if value not in (None, ""):
            return value
    return None


def _write_single_sheet_workbook(
    output_path: Path,
    dataframe: pd.DataFrame,
    sheet_name: str,
) -> None:
    """Write a one-sheet workbook.

    Args:
        output_path: Output workbook path.
        dataframe: Data to write.
        sheet_name: Worksheet name.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(list(dataframe.columns))
    for row in dataframe.itertuples(index=False, name=None):
        sheet.append(["" if value is None else value for value in row])
    workbook.save(output_path)


def _write_multi_sheet_workbook(
    output_path: Path,
    sheets: dict[str, pd.DataFrame],
) -> None:
    """Write a multi-sheet workbook from a mapping of name to DataFrame.

    Args:
        output_path: Output workbook path.
        sheets: Mapping of worksheet name to DataFrame.
    """
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for sheet_name, dataframe in sheets.items():
        sheet = workbook.create_sheet(title=sheet_name)
        sheet.append(list(dataframe.columns))
        for row in dataframe.itertuples(index=False, name=None):
            sheet.append(["" if value is None else value for value in row])

    workbook.save(output_path)
