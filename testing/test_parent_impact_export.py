"""Parent-impact export tests."""

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from arcgis_dependency import interrogate_parent_dependency_impact


@dataclass
class FakeItem:
    """Minimal item stand-in for dependency traversal tests."""

    item_id: str
    title: str
    dependencies: list[object] = field(default_factory=list)

    @property
    def id(self) -> str:
        return self.item_id

    def get_dependencies(
        self,
        deep: bool = False,
        outside_org: bool = False,
        out_format: str = "item",
    ) -> list[object]:
        return list(self.dependencies)


class FakeGISContent:
    """Minimal GIS content manager."""

    def __init__(self, items: dict[str, FakeItem | None]) -> None:
        self._items = items

    def get(self, item_id: str) -> FakeItem | None:
        return self._items.get(item_id)

    def search(self, query: str = "*", max_items: int = 10000) -> list[FakeItem]:
        return [item for item in self._items.values() if item is not None]


class FakeGIS:
    """Minimal GIS wrapper for export tests."""

    def __init__(self, items: dict[str, FakeItem | None]) -> None:
        self.content = FakeGISContent(items)


def test_parent_impact_writes_dual_sheet_workbook_and_returns_path(temp_dir) -> None:
    """Export mode should write impact_rows and target_decisions sheets."""
    target = FakeItem("target", "Target")
    app = FakeItem("app", "App", dependencies=["target"])
    gis = FakeGIS({"target": target, "app": app})

    output_path = temp_dir / "parent-impact.xlsx"
    result = interrogate_parent_dependency_impact(
        target_item_ids=["target"],
        gis=gis,
        output_excel=output_path,
    )

    assert isinstance(result, Path)
    assert result == output_path
    assert result.exists()

    workbook = load_workbook(output_path)
    assert "impact_rows" in workbook.sheetnames
    assert "target_decisions" in workbook.sheetnames

    impact_values = [
        value
        for row in workbook["impact_rows"].iter_rows(values_only=True)
        for value in row
        if value is not None
    ]
    decision_values = [
        value
        for row in workbook["target_decisions"].iter_rows(values_only=True)
        for value in row
        if value is not None
    ]

    assert "target" in impact_values
    assert any(str(value) in {"safe_to_delete", "not_safe_to_delete", "unknown_requires_review"} for value in decision_values)
